import openpyxl


def get_data(sheetName):
    workbook = openpyxl.load_workbook("..//TestData//LoginExcel.xlsx")
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    print("total cols are ", str(total_cols))
    print("total rows are ", str(total_rows))
    mainList = []

    for i in range(2, total_rows + 1):
        dataList = []
        for j in range(1, total_cols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
        print(mainList)
    return mainList
