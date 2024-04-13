import openpyxl

path = "../TestData/LoginExcel.xlsx"


def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName, row, col):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=row, column=col).value


def getAllCellData(sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    total_row = sheet.max_row
    total_column = sheet.max_column
    newList = []

    for i in range(2, total_row + 1):
        dataList = []
        for j in range(1, total_column + 1):
            cell = sheet.cell(row=i, column=j).value
            dataList.insert(j, cell)
        newList.insert(i, dataList)
    return newList


def setCellData(path, sheetName, row, col, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=row, column=col).value = data
    workbook.save(path)

#
# print(getCellData(path, "MobileNumber", 2, 1))
# print(getAllCellData(path, "MobileNumber", 2, 1))
# print(setCellData(path, "MobileNumber", 1, 2, data="Name"))
